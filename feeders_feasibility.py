import py_dss_interface
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time
import os
import pandas as pd
import feeders_list_all


limiar = 15 #Usuario define o limiar de diferença percentual entre a energia do medidor e a energia do BDGD

# Obter o diretório onde o script está localizado
script_dir = os.path.dirname(os.path.abspath(__file__))

# Carregar o arquivo CSV
ctmt = pd.read_csv(r'Inputs\CTMT_Uberlandia.csv', sep=',')

feeders_list = feeders_list_all.get_cod_id_list()
# Inicializa a interface do PyDSS
dss = py_dss_interface.DSSDLL()

# Lista para armazenar os dados de cada feeder
data_list = []

def montar_planilha():
    time_start = time.time()
    for feeder in feeders_list:
        dss.file = rf"C:\FeedersUdia\{feeder}\Master_{feeder}.dss"
        dss.text(f"compile {dss.file}")

        # Resolve o circuito
        dss.solution_solve()

        energymeterday = 0
        energymeterday = dss.meters_register_values()[0]
    
        energymeteryear = 365 * energymeterday

        # Filtra a tabela para o feeder específico
        filtered_ctmt = ctmt[ctmt['COD_ID'] == feeder]

        # Soma das colunas de energia (ENE_01 até ENE_12)
        energy_columns = [f'ENE_{str(i).zfill(2)}' for i in range(1, 13)]
        energyBDGD = filtered_ctmt[energy_columns].sum().sum()
        perc = abs((energymeteryear / energyBDGD) - 1) * 100

        teste_conv = dss.solution_read_converged()  # Verifica se convergiu (1) ou não (0)
        if teste_conv == 0 or energymeteryear == 0 or energymeteryear is None:
            conv = 'Não'
            energymeteryear = 'N/A'
            perc = 'N/A'  # Valor alto para destacar que não convergiu
        else:
            conv = 'Sim'

        # Adicionar os dados na lista
        data_list.append({
            "Feeder": feeder,
            "Converged": conv,
            "Energy Meter Year": energymeteryear,
            "Energy BDGD": energyBDGD,
            "Percent Difference": perc
        })

    # Criar o DataFrame final com todos os feeders
    df = pd.DataFrame(data_list)

    # Salvar o DataFrame em uma planilha Excel
    filename = os.path.join(script_dir, "ComparativoFeeders.xlsx")
    df.to_excel(filename, index=False)

    for index, row in df.iterrows():
        if pd.isna(row["Energy Meter Year"]) or row["Energy Meter Year"] == 0:
            df.at[index, "Converged"] = "Não"
            df.at[index, "Energy Meter Year"] = "N/A"
            df.at[index, "Percent Difference"] = "N/A"

    # Salvar as alterações na planilha
    df.to_excel(filename, index=False)

    # Carregar a planilha para aplicar formatação
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Definindo o estilo de preenchimento (cor)
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Aplicar formatação: pintar células na coluna 'Percent Difference'
    for row in range(2, sheet.max_row + 1):  # Começar da linha 2 para evitar o cabeçalho
        cell = sheet[f"E{row}"]  # Coluna D contém 'Percent Difference'
        if cell.value is not None and cell.value != "N/A":  # Verifica se o valor não é None nem 'N/A'
            try:
                numeric_value = float(cell.value)
                if numeric_value > limiar:  
                    cell.fill = fill_red
                else:  
                    cell.fill = fill_green
            except ValueError:
                pass  # Caso o valor não possa ser convertido para float, não aplicar formatação
        if cell.value == "N/A":
            cell.fill = fill_red

    # Salvar as alterações no arquivo
    workbook.save(filename)

    time_end = time.time()
    print(f"ComparativoFeeders.xlsx criada e formatada com sucesso em {time_end-time_start}!")

montar_planilha()